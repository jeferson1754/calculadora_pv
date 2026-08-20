import win32com.client
import sys
import subprocess
import time
import config_sap
import io
import pandas as pd
import os
import pyperclip
from sqlalchemy import create_engine
import pymysql
from pathlib import Path
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from auditoria import RegistrarEjecucion


ruta = r"C:\xampp\htdocs\Calculadora_PV\PV"

# Convertir la ruta a un objeto Path
base_dir = Path(r"C:\xampp\htdocs\Calculadora_PV\PV")
base_bom = Path(r"C:\xampp\htdocs\Calculadora_PV\PV\BOM")

archivo_origen = base_dir / "BASE_PEDIDOS_ORIGINAL.XLS"
# Ahora el operador / funciona perfectamente
archivo_txt = base_dir / "CODIGOS_SAP_UNICOS.txt"
archivo_xlsx = base_dir / "CODIGOS_SAP_UNICOS.xlsx"

salida_trabajo = base_dir / "BASE_TRABAJO.xlsx"
salida_revision_37 = base_dir / "BASE_REVISION_37.xlsx"
salida_productiva = base_dir / "BASE_PRODUCTIVA.xlsx"
salida_unicos_txt = base_dir / "CODIGOS_SAP_UNICOS.txt"
salida_unicos_excel = base_dir / "CODIGOS_SAP_UNICOS.xlsx"

salida_bom = base_dir / "BASE_BOM.xlsx"
archivo_bom_filtrada = base_dir / "BASE_BOM_FILTRADA.xlsx"
archivo_salida = base_dir / "BASE_PRODUCTIVA_BOM.xlsx"

salida_excel_operativo = base_dir / "VISTA_OPERATIVA_PEDIDOS_BOM.xlsx"

columnas_requeridas = [
    "Nombre",
    "Pedido",
    "PosPed",
    "Valor neto",
    "Mon.",
    "Ctd.ped.",
    "Material",
    "Fecha Probable",
]

# --- CONFIGURACIÓN DE CONEXIÓN MYSQL (XAMPP) ---
# Ajusta el usuario, contraseña y puerto según tu configuración de XAMPP
USER_DB = "root"
PASSWORD_DB = ""  # Por defecto en XAMPP la clave viene vacía
HOST_DB = "localhost"
PORT_DB = "3306"
NOMBRE_DB = "calculadora_pv"
CANTIDAD_CODIGOS = 50


# URL de conexión SQLAlchemy
cadena_conexion = f"mysql+pymysql://{USER_DB}:{PASSWORD_DB}@{HOST_DB}:{PORT_DB}/{NOMBRE_DB}"

def saplogin():
    try:
        path = r"C:\Program Files (x86)\SAP\FrontEnd\SAPgui\saplogon.exe"
        subprocess.Popen(path)
        time.sleep(1)
        SapGuiAuto = win32com.client.GetObject('SAPGUI')
        if not type(SapGuiAuto) == win32com.client.CDispatch:
            return

        application = SapGuiAuto.GetScriptingEngine
        if not type(application) == win32com.client.CDispatch:
            SapGuiAuto = None
            return

        # connection = application.OpenConnection("PRD [REVESOL]",True)
        connection = application.OpenConnection("SR  [SAP ROUTER]", True)
        session = connection.Children(0)
        if not type(session) == win32com.client.CDispatch:
            connection = None
            application = None
            SapGuiAuto = None
            return

    except:
        print(sys.exc_info())
        print(4)

    session.findById("wnd[0]/usr/txtRSYST-BNAME").text = config_sap.username
    session.findById("wnd[0]/usr/pwdRSYST-BCODE").text = config_sap.password
    session.findById("wnd[0]").sendVKey(0)
    session.findById("wnd[0]").sendVKey(0)
    # session.findById("wnd[0]").resizeWorkingPane(98,16,False)
    return session

def extraccion_pendientes(fecha_desde, fecha_hasta,session, path=ruta,
    nombre_archivo="BASE_PEDIDOS_ORIGINAL.XLS"):

    session.findById("wnd[0]/usr/ctxtS_ERDAT-LOW").text = fecha_desde
    session.findById("wnd[0]/usr/ctxtS_ERDAT-HIGH").text = fecha_hasta
    
    time.sleep(5)
    session.findById("wnd[0]/tbar[1]/btn[8]").press()

    time.sleep(5)
    # LEE PEDIDOS PENDIENTES    
    print(f"Se procede a descargar la OF {nombre_archivo}")
    session.findById("wnd[0]").maximize()
    session.findById("wnd[0]/tbar[1]/btn[45]").press()
    session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").selected=True
    session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").setFocus()
    session.findById("wnd[1]/tbar[0]/btn[0]").press()


    # 4. Guardado de archivo pasando la variable de ruta correctamente
    # Asegurar que el path no falle en SAP GUI
    path_sap = os.path.normpath(path)
    #GUARDADO DE ARCHIVO
    session.findById("wnd[1]/usr/ctxtDY_PATH").text = path_sap
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = nombre_archivo

    session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 11
    # session.findById("wnd[1]").sendVKey(0)
    session.findById("wnd[1]/tbar[0]/btn[11]").press()

    time.sleep(10)

    session.findById("wnd[0]/tbar[0]/btn[3]").press()

    session.findById("wnd[0]").maximize


    print(f"Se descargó la OF {nombre_archivo} con éxito en la ruta {ruta}")

    session.findById("wnd[0]/tbar[0]/btn[3]").press()

def obtener_fechas_sap():
    conexion = pymysql.connect(
        host='localhost',
        user='root',
        password='',
        database='calculadora_pv'
    )
    with conexion.cursor() as cursor:
        cursor.execute("SELECT clave, valor FROM parametros_configuracion WHERE clave IN ('FECHA_DESDE', 'FECHA_HASTA')")
        filas = dict(cursor.fetchall())
        
    conexion.close()
    return filas.get('FECHA_DESDE', '01.01.2026'), filas.get('FECHA_HASTA', '31.12.2026')

def dividir_lista(lista, tamano_lote):
    """Generador para dividir una lista en sublistas de un tamaño determinado."""
    for i in range(0, len(lista), tamano_lote):
        yield lista[i : i + tamano_lote]

def obtener_codigos_unicos():
    """Lee los códigos SAP únicos desde el TXT, Excel o BASE_PRODUCTIVA."""

    # 1. Intentar desde el TXT
    if archivo_txt.exists():
        with open(archivo_txt, "r", encoding="utf-8") as f:
            codigos = [
                linea.strip()
                for linea in f
                if linea.strip() and not linea.strip().startswith("#")
            ]
        return list(dict.fromkeys(codigos))

    # 2. Intentar desde el Excel de únicos
    if archivo_xlsx.exists():
        df = pd.read_excel(archivo_xlsx, dtype=str)
        col = df.columns[0]
        codigos = df[col].dropna().astype(str).str.strip().tolist()
        return list(dict.fromkeys(codigos))

    raise FileNotFoundError(
        "No se encontró ningún archivo de códigos SAP únicos."
    )

def vaciar_carpeta(ruta_carpeta):
    """
    Elimina todos los archivos y subcarpetas dentro de la ruta especificada,
    manteniendo la carpeta principal intacta.
    """
    carpeta = Path(ruta_carpeta)

    # Verificar que la carpeta exista
    if not carpeta.exists():
        print(f"[ADVERTENCIA]: La carpeta '{ruta_carpeta}' no existe.")
        return False

    if not carpeta.is_dir():
        print(f"[ERROR]: La ruta '{ruta_carpeta}' no es un directorio válido.")
        return False

    archivos_eliminados = 0
    carpetas_eliminadas = 0

    try:
        # Iterar sobre todo el contenido de la carpeta
        for elemento in carpeta.iterdir():
            if elemento.is_file() or elemento.is_symlink():
                elemento.unlink()  # Elimina archivos o enlaces
                archivos_eliminados += 1
            elif elemento.is_dir():
                shutil.rmtree(elemento)  # Elimina subcarpetas y su contenido
                carpetas_eliminadas += 1

        print(f"¡Éxito! Se limpió la carpeta '{carpeta.name}':")
        print(f" - Archivos borrados: {archivos_eliminados}")
        print(f" - Subcarpetas borradas: {carpetas_eliminadas}")
        return True

    except Exception as e:
        print(f"[ERROR] Ocurrió un problema al vaciar la carpeta: {e}")
        return False

def procesar_descarga_por_lotes(session):
    # Límite por lote para el portapapeles de SAP GUI
    """Obtiene los códigos únicos y los procesa en lotes de 60 hacia SAP GUI."""

    lista_of_madres = obtener_codigos_unicos()
    total_codigos = len(lista_of_madres)

    if not total_codigos:
        print("ADVERTENCIA: La lista de códigos está vacía.")
        return

    print(f"\nTotal de códigos a procesar: {total_codigos}")
    print(f"Tamaño de lote configurado: {CANTIDAD_CODIGOS}")
    print("--------------------------------------------------")
        
    vaciar_carpeta(base_bom)



    for numero_lote, lote in enumerate(
        dividir_lista(lista_of_madres, CANTIDAD_CODIGOS), start=1
    ):
        
        texto_portapapeles = "\r\n".join(str(total_codigos).strip() for total_codigos in lote)

        # Copiar al portapapeles de Windows
        pyperclip.copy(texto_portapapeles)

        nombre_archivo_lote = f"BOM_{numero_lote:03d}.XLS"
        print(
            f"Procesando lote {numero_lote:03d} de {len(lote)} OFs -> {nombre_archivo_lote}"
        )

        
        # ENTRAMOS A LA TRANSACCIÓN
        session.findById("wnd[0]").maximize()
        session.findById("wnd[0]/tbar[0]/okcd").text = "ZMM_0002"
        session.findById("wnd[0]").sendVKey(0)

        # Llamar a la función que interactúa con la sesión de SAP GUI
        descargar_materiales(
            texto_ofs=texto_portapapeles,
            path=str(base_bom),
            session=session,
            nombre_archivo=nombre_archivo_lote,
        )
        
        session.findById("wnd[0]/tbar[0]/btn[3]").press()   

def descargar_materiales(texto_ofs, path, session, nombre_archivo):


    # INGRESA A LA LISTA PARA COPIAR LAS OFS y SALIR DE LA LISTA
    session.findById("wnd[0]/usr/btn%_S_MATNR_%_APP_%-VALU_PUSH").press()
    session.findById("wnd[1]/tbar[0]/btn[24]").press()
    session.findById("wnd[1]/tbar[0]/btn[8]").press()
    time.sleep(5)
    session.findById("wnd[0]/tbar[1]/btn[8]").press()
    time.sleep(5)

    print(f"Se procede a descargar la OF {nombre_archivo}")
    session.findById("wnd[0]").maximize
    session.findById("wnd[0]/tbar[1]/btn[45]").press()

    session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").selected = True
    session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").setFocus

    session.findById("wnd[1]/tbar[0]/btn[0]").press()


    #GUARDADO DE ARCHIVO
    session.findById("wnd[1]/usr/ctxtDY_PATH").text = path
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = nombre_archivo
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 1
    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    session.findById("wnd[0]/tbar[0]/btn[3]").press()

    print(f"Se descargo la OF {nombre_archivo}")
    # 3. Función de limpieza estricta

def limpiar_texto(val):
    if pd.isna(val):
        return ""
    val_str = str(val).strip()
    if val_str.endswith(".0"):
        val_str = val_str[:-2]
    return val_str

def procesar_excel_productivo_1(columnas_requeridas=columnas_requeridas, archivo_origen=archivo_origen,salida_trabajo=salida_trabajo, salida_revision_37=salida_revision_37,salida_productiva=salida_productiva,salida_unicos_txt=salida_unicos_txt,salida_unicos_excel=salida_unicos_excel):
# 1. Encontrar la fila de encabezados con detección flexible de codificación
# 1. Encontrar la fila de encabezados
    header_row = None
    encoding_exitoso = None

    # Probar codificaciones típicas de SAP
    for enc in ["utf-16", "latin1", "utf-8", "cp1252"]:
        try:
            with open(archivo_origen, "r", encoding=enc, errors="ignore") as f:
                for i, line in enumerate(f):
                    linea_limpia = line.strip()
                    # Buscar columnas clave exactas según la salida de SAP
                    if (
                        "Doc.venta" in linea_limpia
                        or "Nº ped.cliente" in linea_limpia
                    ) and "Material" in linea_limpia:
                        header_row = i
                        encoding_exitoso = enc
                        break
            if header_row is not None:
                break
        except Exception:
            continue

    if header_row is None:
        raise ValueError(
            f"No se pudo detectar la fila de encabezados en el archivo '{archivo_origen}'."
        )

    # 2. Leer el archivo saltando hasta la fila del encabezado encontrada
    df = pd.read_csv(
        archivo_origen,
        sep="\t",
        encoding=encoding_exitoso,
        skiprows=header_row,
        engine="python",
        on_bad_lines="skip",
        dtype=str,
    )

    # Limpiar espacios en los nombres de las columnas
    df.columns = [str(c).strip() for c in df.columns]

    # Si en el script utilizas 'Pedido' como nombre de columna,
    # renombramos 'Doc.venta' a 'Pedido' para mantener compatibilidad
    if "Doc.venta" in df.columns and "Pedido" not in df.columns:
        df.rename(columns={"Doc.venta": "Pedido"}, inplace=True)
    # Limpiar nombres de columnas
    df.columns = df.columns.str.strip()

    # Aplicar limpieza a todas las columnas
    for col in df.columns:
        df[col] = df[col].apply(limpiar_texto)

    df_trabajo = df[columnas_requeridas].copy()

    # Guardar BASE_TRABAJO (Aquí SÍ se repiten los materiales porque son múltiples pedidos)
    df_trabajo.to_excel(salida_trabajo, index=False)

    # =========================================================================
    # SEPARACIÓN DE MATERIALES 37xx
    # =========================================================================
    es_37 = df_trabajo["Material"].str.startswith("37")

    base_revision_37 = df_trabajo[es_37].copy()
    base_productiva = df_trabajo[~es_37].copy()

    base_revision_37["Estado"] = "REVISIÓN MATERIAL 37"
    base_productiva["Estado"] = "PRODUCTIVO"

    # Guardar bases de pedidos (Aquí SÍ se repiten los materiales)
    base_revision_37.to_excel(salida_revision_37, index=False)
    base_productiva.to_excel(salida_productiva, index=False)

    # =========================================================================
    # OBTENCIÓN DE LA LISTA DE CÓDIGOS SAP ÚNICOS (SIN REPETIDOS)
    # =========================================================================
    # Extraemos la serie de materiales, quitamos vacíos y eliminamos duplicados
    codigos_unicos_series = (
        base_productiva["Material"]
        .replace("", pd.NA)
        .dropna()
        .drop_duplicates()
        .sort_values()
    )

    # Convertir a lista de Python
    codigos_sap_unicos = codigos_unicos_series.tolist()

    # Opción A: Guardar en archivo TXT (Ideal para copiar y pegar en SAP ZMM_0002)
    with open(salida_unicos_txt, "w", encoding="utf-8") as f:
        for codigo in codigos_sap_unicos:
            f.write(f"{codigo}\n")

    # Opción B: Guardar en Excel de códigos únicos sin duplicados
    df_unicos = pd.DataFrame({"Material_Unico": codigos_sap_unicos})
    df_unicos.to_excel(salida_unicos_excel, index=False)

    print("\n--- RESUMEN DE PROCESAMIENTO ---")
    print(f"1. Total registros en BASE_TRABAJO: {len(df_trabajo)}")
    print(f"2. Total registros en BASE_PRODUCTIVA: {len(base_productiva)}")
    print(
        f"3. CÓDIGOS ÚNICOS PARA ZMM_0002: {len(codigos_sap_unicos)} (¡Sin duplicados!)"
    )
    print(f"\nArchivo TXT listo para SAP: {salida_unicos_txt}")
    print(f"Archivo Excel de únicos: {salida_unicos_excel}")

def filtrar_columnas_bom():
    if not salida_bom.exists():
        raise FileNotFoundError(
            f"No se encontró el archivo origen: {salida_bom}"
        )

    print(f"Leyendo archivo consolidado: {salida_bom.name}...")
    df_bom = pd.read_excel(salida_bom, dtype=str)

    # Limpiar espacios en los nombres de las columnas
    df_bom.columns = df_bom.columns.astype(str).str.strip()

    print("\nColumnas disponibles en el archivo cargado:")
    print(list(df_bom.columns))

    # Mapeo flexible para asegurar captura de Cantidad y UMB
    alias_columnas = {
        "Material": "Material",
        "Texto breve de material": "Texto breve de material",
        "Nivel Explosión": "Nivel Explosión",
        "Pos.": "Pos.",
        "N° Componentes": "N° Componentes",
        "Desc. Componente": "Desc. Componente",
        "Cantidad": [c for c in df_bom.columns if "Cantidad" in c][0]
        if any("Cantidad" in c for c in df_bom.columns)
        else "Cantidad",
        "UMB": [c for c in df_bom.columns if "UMB" in c][0]
        if any("UMB" in c for c in df_bom.columns)
        else "UMB",
        "Ce.": [c for c in df_bom.columns if "Ce." in c][0]
        if any("Ce." in c for c in df_bom.columns)
        else "Ce.",
        "Alm.": [c for c in df_bom.columns if "Alm." in c][0]
        if any("Alm." in c for c in df_bom.columns)
        else "Alm.",
        
    }

    # Filtrar solo las columnas existentes
    cols_existentes = [
        col for col in alias_columnas.values() if col in df_bom.columns
    ]

    if not cols_existentes:
        raise ValueError(
            "No se pudieron hacer coincidir las columnas solicitadas."
        )

    df_filtrado = df_bom[cols_existentes].copy()

    # Renombrar a los nombres estándar finales
    renombres = {v: k for k, v in alias_columnas.items() if v in df_bom.columns}
    df_filtrado.rename(columns=renombres, inplace=True)

    # Limpieza de espacios en celdas de datos
    for col in df_filtrado.columns:
        df_filtrado[col] = df_filtrado[col].fillna("").astype(str).str.strip()

    # Guardar archivo filtrado definitivo
    df_filtrado.to_excel(archivo_bom_filtrada, index=False)

    print("\n==================================================")
    print(
        f"¡ÉXITO! Se filtraron las {len(df_filtrado.columns)} columnas requeridas con {len(df_filtrado)} registros."
    )
    print("Columnas finales en el archivo:")
    print(list(df_filtrado.columns))
    print(f"\nArchivo final guardado en:\n{salida_bom}")
    print("==================================================")

def detectar_encabezado_y_codificacion(ruta_archivo):
    """Prueba codificaciones para hallar la fila donde comienza la tabla del BOM."""
    codificaciones = ["utf-16", "latin1", "utf-8", "cp1252"]
    
    for enc in codificaciones:
        try:
            with open(ruta_archivo, "r", encoding=enc, errors="ignore") as f:
                for i, line in enumerate(f):
                    # Buscamos 'Material' y algún otro campo característico de la BOM
                    if "Material" in line and ("Cantidad" in line or "Pos" in line or "UMB" in line):
                        return i, enc
        except Exception:
            continue
            
    return None, None

def cargar_archivo_bom_lote(ruta_archivo):
    """Lee un lote individual de BOM detectando dinámicamente su formato."""
    
    header_row, encoding_detectado = detectar_encabezado_y_codificacion(ruta_archivo)

    if header_row is None:
        return None

    try:
        df = pd.read_csv(
            ruta_archivo,
            sep="\t",
            encoding=encoding_detectado,
            skiprows=header_row,
            engine="python",
            on_bad_lines="skip",
            dtype=str
        )

        # Limpiar espacios en blanco de nombres de columnas
        df.columns = df.columns.astype(str).str.strip()
        return df
    except Exception as e:
        print(f"  [Error al leer {ruta_archivo.name}]: {e}")
        return None

def procesar_todos_los_lotes_bom(base_bom):
    if not base_bom.exists():
        raise FileNotFoundError(f"No existe la carpeta de descargas: {base_bom}")

    # Buscar tanto BOM_*.XLS como MATERIALES_LOTE_*.XLS
    archivos_lotes = sorted(list(base_bom.glob("*.XLS")) + list(base_bom.glob("*.xls")))

    if not archivos_lotes:
        raise FileNotFoundError(f"No se encontraron archivos .XLS en {base_bom}")

    dataframes_lotes = []

    for archivo in archivos_lotes:
        print(f"Procesando: {archivo.name}...")
        df_lote = cargar_archivo_bom_lote(archivo)

        if df_lote is None or df_lote.empty:
            print(f"  [Omitido] No se hallaron encabezados válidos en {archivo.name}")
            continue

        # Filtrar solo columnas válidas que contengan datos reales (excluir columnas vacías creadas por tabuladores extra)
        cols_validas = [c for c in df_lote.columns if c and not c.startswith("Unnamed")]
        df_filtrado = df_lote[cols_validas].copy()

        dataframes_lotes.append(df_filtrado)

    if not dataframes_lotes:
        raise Exception("No se pudo extraer información válida de ningún lote.")

    # Consolidar todos los DataFrames
    base_bom = pd.concat(dataframes_lotes, ignore_index=True)

    # Limpiar espacios en blanco de todas las celdas
    for col in base_bom.columns:
        base_bom[col] = base_bom[col].astype(str).str.strip()

    # Eliminar filas repetidas que coincidan con la fila de encabezados
    primer_col = base_bom.columns[0]
    base_bom = base_bom[base_bom[primer_col] != primer_col].copy()

    # Guardar BASE_BOM.xlsx
    base_bom.to_excel(salida_bom, index=False)

    print("\n==================================================")
    print(f"¡ÉXITO! BASE_BOM generada con {len(base_bom)} filas de componentes.")
    print(f"Columnas detectadas en la BASE_BOM:")
    print(list(base_bom.columns[:10]))  # Muestra las primeras 10 columnas
    print(f"Archivo guardado en: {salida_bom}")
    print("==================================================")
    filtrar_columnas_bom()

def a_numero(val):
    """Convierte cadenas de texto numéricas de SAP (con coma o punto) a float."""
    if not val or str(val).lower() == "nan":
        return 0.0
    val_clean = str(val).strip()
    if "," in val_clean:
        val_clean = val_clean.replace(".", "").replace(",", ".")
    try:
        return float(val_clean)
    except ValueError:
        return 0.0

def vincular_bases(archivo_productiva=salida_productiva, archivo_bom=salida_bom):
    if not archivo_productiva.exists():
        raise FileNotFoundError(f"No se encontró el archivo: {archivo_productiva}")
    if not archivo_bom.exists():
        raise FileNotFoundError(f"No se encontró el archivo: {archivo_bom}")

    print("Cargando BASE_PRODUCTIVA y BASE_BOM...")
    df_prod = pd.read_excel(archivo_productiva, dtype=str)
    df_bom = pd.read_excel(archivo_bom, dtype=str)

    # 1. Limpieza de nombres de columnas y celdas
    df_prod.columns = df_prod.columns.str.strip()
    df_bom.columns = df_bom.columns.str.strip()

    for col in df_prod.columns:
        df_prod[col] = df_prod[col].fillna("").astype(str).str.strip()
    for col in df_bom.columns:
        df_bom[col] = df_bom[col].fillna("").astype(str).str.strip()

    df_prod.drop_duplicates(
        subset=["Pedido", "PosPed", "Material"], keep="first", inplace=True
    )

    # Eliminar duplicados en BASE_BOM por Producto Padre + Componente + Nivel + Posición BOM
    subset_bom = [
        c
        for c in ["Material", "N° Componentes", "Componente", "Nivel Explosión", "Pos."]
        if c in df_bom.columns
    ]
    df_bom.drop_duplicates(subset=subset_bom, keep="first", inplace=True)

    # 2. Conversión a formato numérico para poder multiplicar
    df_prod["Ctd_Ped_Num"] = df_prod["Ctd.ped."].apply(a_numero)
    df_bom["Cantidad_BOM_Num"] = df_bom["Cantidad"].apply(a_numero)

    # 3. Cruce/Merge por el Código SAP de Material
    # BASE_PRODUCTIVA (Material) <---> BASE_BOM (Material)
    print("Realizando el cruce entre BASE_PRODUCTIVA y BASE_BOM...")
    df_cruce = pd.merge(
        df_prod, df_bom, on="Material", how="inner", suffixes=("_PD", "_BOM")
    )

    # 4. Cálculo: Cantidad Total Requerida del Componente por Pedido
    # Cantidad Total = (Cantidad Pedida del Producto) * (Cantidad Unitaria del Componente en BOM)
    df_cruce["Cantidad_Total_Requerida"] = (
        df_cruce["Ctd_Ped_Num"] * df_cruce["Cantidad_BOM_Num"]
    ).round(4)

    # 5. Eliminar columnas auxiliares numéricas
    df_cruce.drop(columns=["Ctd_Ped_Num", "Cantidad_BOM_Num"], inplace=True)

    # DEPURACIÓN FINAL EN LA BASE VINCULADA
    subset_final = ["Pedido", "PosPed", "Material", "N° Componentes", "Nivel Explosión"]
    subset_existentes = [c for c in subset_final if c in df_cruce.columns]
    df_cruce.drop_duplicates(subset=subset_existentes, keep="first", inplace=True)

    # Guardar
    df_cruce.to_excel(archivo_salida, index=False)
    print(f"¡Éxito! Archivo guardado con {len(df_cruce)} filas únicas en:\n{archivo_salida}")

    print("\n==================================================")
    print(
        f"¡ÉXITO! Vinculación completada. Se generaron {len(df_cruce)} filas cruzadas."
    )
    print(f"Archivo resultante guardado en: {archivo_salida}")
    print("==================================================")

    return len(df_cruce)

def guardar_en_mysql(archivo_trabajo=salida_trabajo, archivo_rev37=salida_revision_37, archivo_productiva=salida_productiva, archivo_bom=salida_bom, archivo_vinculada=archivo_salida):
    """Carga los DataFrames y los guarda como tablas en MySQL."""
    print("Iniciando conexión con MySQL (XAMPP)...")
    engine = create_engine(cadena_conexion)

    # 1. Cargar DataFrames
    print("Leyendo archivos Excel...")
    df_trabajo = pd.read_excel(archivo_trabajo, dtype=str) if archivo_trabajo.exists() else None
    df_rev37 = pd.read_excel(archivo_rev37, dtype=str) if archivo_rev37.exists() else None
    df_productiva = pd.read_excel(archivo_productiva, dtype=str) if archivo_productiva.exists() else None
    df_bom = pd.read_excel(archivo_bom, dtype=str) if archivo_bom.exists() else None
    df_vinculada = pd.read_excel(archivo_vinculada, dtype=str) if archivo_vinculada.exists() else None

    # 2. Insertar en MySQL
    with engine.begin() as conn:
        if df_trabajo is not None:
            df_trabajo.to_sql("pedidos_pendientes_todos", con=conn, if_exists="replace", index=False)
            print("  -> Tabla 'pedidos_pendientes_todos' guardada en MySQL.")

        if df_rev37 is not None:
            df_rev37.to_sql("pedidos_revision_37", con=conn, if_exists="replace", index=False)
            print("  -> Tabla 'pedidos_revision_37' guardada en MySQL.")

        if df_productiva is not None:
            df_productiva.to_sql("pedidos_productivos", con=conn, if_exists="replace", index=False)
            print("  -> Tabla 'pedidos_productivos' guardada en MySQL.")

        if df_bom is not None:
            df_bom.to_sql("base_bom_componentes", con=conn, if_exists="replace", index=False)
            print("  -> Tabla 'base_bom_componentes' guardada en MySQL.")

        if df_vinculada is not None:
            df_vinculada.to_sql("vista_productiva_bom", con=conn, if_exists="replace", index=False)
            print("  -> Tabla 'vista_productiva_bom' guardada en MySQL.")

    print("¡Persistencia en MySQL completada con éxito!\n")
    return df_vinculada, df_rev37

def generar_excel_operativo(df_vinculada, df_rev37):
    """Genera un archivo Excel de trabajo profesional unificado para los usuarios."""
    print("Generando Vista Operativa de Excel formateada...")

    # Consolidar ambas vistas para la tabla unificada de trabajo
    dfs_a_unir = []
    if df_vinculada is not None and not df_vinculada.empty:
        df_v = df_vinculada.copy()
        if "Estado" not in df_v.columns:
            df_v["Estado"] = "PRODUCTIVO"
        dfs_a_unir.append(df_v)

    if df_rev37 is not None and not df_rev37.empty:
        df_r = df_rev37.copy()
        df_r["Estado"] = "REVISIÓN MATERIAL 37"
        dfs_a_unir.append(df_r)

    df_operativo = pd.concat(dfs_a_unir, ignore_index=True)

    # Ordenar columnas estratégicamente
    cols_orden = [
        "Cliente" if "Cliente" in df_operativo.columns else "Nombre",
        "Pedido",
        "PosPed",
        "Material",
        "Ctd.ped.",
        "ValorNeto",
        "Mon.",
        "Fecha Probable",
        "Estado",
        "Nivel Explosión",
        "Pos.",
        "N° Componentes",
        "Desc. Componente",
        "Cantidad",
        "UMB",
        "Cantidad_Total_Requerida"
    ]

    # Conservar solo las columnas existentes
    cols_finales = [c for c in cols_orden if c in df_operativo.columns]
    df_operativo = df_operativo[cols_finales]

    # Crear libro con openpyxl para estilo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vista Operativa"

    # Estilos de encabezado y bordes
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Escribir Cabecera
    headers = list(df_operativo.columns)
    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Escribir Datos
    for row in df_operativo.itertuples(index=False):
        ws.append(list(row))

    # Formatear filas y celdas
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border_thin
            col_name = headers[col_idx - 1]

            # Formatos numéricos y alineación
            if col_name in ["Pedido", "PosPed", "Material", "N° Componentes"]:
                cell.alignment = Alignment(horizontal="center")
            elif col_name in ["Ctd.ped.", "Cantidad", "Cantidad_Total_Requerida"]:
                cell.alignment = Alignment(horizontal="right")
            elif col_name == "ValorNeto":
                cell.alignment = Alignment(horizontal="right")
            elif col_name == "Estado":
                cell.alignment = Alignment(horizontal="center")
                if cell.value == "REVISIÓN MATERIAL 37":
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                elif cell.value == "PRODUCTIVO":
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # Activar Filtro Automático en Excel
    ws.auto_filter.ref = ws.dimensions

    # Autoajustar ancho de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(salida_excel_operativo)
    print(f"¡Éxito! Vista Operativa de Excel generada en:\n{salida_excel_operativo}")


# --- PUNTO DE ENTRADA ---
if __name__ == "__main__":
    
    # --- EJECUCIÓN PRINCIPAL CON REGISTRO DE AUDITORÍA ---
    with RegistrarEjecucion(modulo="Carga Masiva BOM y Pedidos") as ejecucion:

        session_sap = saplogin()

        vaciar_carpeta(base_dir)

        session_sap.findById("wnd[0]").maximize()
        session_sap.findById("wnd[0]/tbar[0]/okcd").text = "ZSD_0001"
        session_sap.findById("wnd[0]").sendVKey(0)
        # Uso en tu automatización SAP GUI:
        fecha_inicio, fecha_fin = obtener_fechas_sap()

        extraccion_pendientes(fecha_inicio, fecha_fin, session_sap)

        procesar_excel_productivo_1()

        procesar_descarga_por_lotes(session_sap)

        try:
            procesar_todos_los_lotes_bom(base_bom)
        except Exception as e:
            print(f"\n[ERROR]: {e}")

        total_filas = 0

        try:
            total_filas = vincular_bases()
        except Exception as e:
            print(f"\n[ERROR]: {e}")

        try:
            df_v, df_r = guardar_en_mysql()
            generar_excel_operativo(df_v, df_r)
        except Exception as e:
            print(f"\n[ERROR]: {e}") 

            # 1. Tu código habitual de procesamiento
        print("Procesando archivos...")

            # 2. Indicar que terminó con éxito pasando el conteo de filas
        ejecucion.finalizar(registros_procesados=total_filas, estado="EXITOSO")

 