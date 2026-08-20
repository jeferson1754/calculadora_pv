from pathlib import Path
import pandas as pd
from sqlalchemy import create_engine
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- RUTAS DE TRABAJO ---
base_dir = Path(r"C:\xampp\htdocs\Calculadora_PV\PV")

archivo_trabajo = base_dir / "BASE_TRABAJO.xlsx"
archivo_rev37 = base_dir / "BASE_REVISION_37.xlsx"
archivo_productiva = base_dir / "BASE_PRODUCTIVA.xlsx"
archivo_bom = base_dir / "BASE_BOM.xlsx"
archivo_vinculada = base_dir / "BASE_PRODUCTIVA_BOM.xlsx"

salida_excel_operativo = base_dir / "VISTA_OPERATIVA_PEDIDOS_BOM.xlsx"

# --- CONFIGURACIÓN DE CONEXIÓN MYSQL (XAMPP) ---
# Ajusta el usuario, contraseña y puerto según tu configuración de XAMPP
USER_DB = "root"
PASSWORD_DB = ""  # Por defecto en XAMPP la clave viene vacía
HOST_DB = "localhost"
PORT_DB = "3306"
NOMBRE_DB = "calculadora_pv"

# URL de conexión SQLAlchemy
cadena_conexion = f"mysql+pymysql://{USER_DB}:{PASSWORD_DB}@{HOST_DB}:{PORT_DB}/{NOMBRE_DB}"


def guardar_en_mysql():
    """Carga los DataFrames y los guarda como tablas en MySQL."""
    print("Iniciando conexión con MySQL (XAMPP)...")
    engine = create_engine(cadena_conexion)

    # 1. Cargar DataFrames
    print("Leyendo archivos Excel...")
    df_trabajo = pd.read_excel(archivo_trabajo, dtype=str) if archivo_trabajo.exists() else None
    df_rev37 = pd.read_excel(archivo_rev37, dtype=str) if archivo_rev37.exists() else None
    df_productiva = pd.read_excel(archivo_productiva, dtype=str) if archivo_productiva.exists() else None
    df_bom = pd.read_excel(archivo_bom, dtype=str) if archivo_bom.exists() else None
    df_vinculada = pd.read_excel(archivo_vinculada, dtype=str) if archivo_vinculada.exists() else None

    # 2. Insertar en MySQL
    with engine.begin() as conn:
        if df_trabajo is not None:
            df_trabajo.to_sql("pedidos_pendientes_todos", con=conn, if_exists="replace", index=False)
            print("  -> Tabla 'pedidos_pendientes_todos' guardada en MySQL.")

        if df_rev37 is not None:
            df_rev37.to_sql("pedidos_revision_37", con=conn, if_exists="replace", index=False)
            print("  -> Tabla 'pedidos_revision_37' guardada en MySQL.")

        if df_productiva is not None:
            df_productiva.to_sql("pedidos_productivos", con=conn, if_exists="replace", index=False)
            print("  -> Tabla 'pedidos_productivos' guardada en MySQL.")

        if df_bom is not None:
            df_bom.to_sql("base_bom_componentes", con=conn, if_exists="replace", index=False)
            print("  -> Tabla 'base_bom_componentes' guardada en MySQL.")

        if df_vinculada is not None:
            df_vinculada.to_sql("vista_productiva_bom", con=conn, if_exists="replace", index=False)
            print("  -> Tabla 'vista_productiva_bom' guardada en MySQL.")

    print("¡Persistencia en MySQL completada con éxito!\n")
    return df_vinculada, df_rev37


def generar_excel_operativo(df_vinculada, df_rev37):
    """Genera un archivo Excel de trabajo profesional unificado para los usuarios."""
    print("Generando Vista Operativa de Excel formateada...")

    # Consolidar ambas vistas para la tabla unificada de trabajo
    dfs_a_unir = []
    if df_vinculada is not None and not df_vinculada.empty:
        df_v = df_vinculada.copy()
        if "Estado" not in df_v.columns:
            df_v["Estado"] = "PRODUCTIVO"
        dfs_a_unir.append(df_v)

    if df_rev37 is not None and not df_rev37.empty:
        df_r = df_rev37.copy()
        df_r["Estado"] = "REVISIÓN MATERIAL 37"
        dfs_a_unir.append(df_r)

    df_operativo = pd.concat(dfs_a_unir, ignore_index=True)

    # Ordenar columnas estratégicamente
    cols_orden = [
        "Cliente" if "Cliente" in df_operativo.columns else "Nombre",
        "Pedido",
        "PosPed",
        "Material",
        "Ctd.ped.",
        "ValorNeto",
        "Mon.",
        "Fecha Probable",
        "Estado",
        "Nivel Explosión",
        "Pos.",
        "N° Componentes",
        "Desc. Componente",
        "Cantidad",
        "UMB",
        "Cantidad_Total_Requerida"
    ]

    # Conservar solo las columnas existentes
    cols_finales = [c for c in cols_orden if c in df_operativo.columns]
    df_operativo = df_operativo[cols_finales]

    # Crear libro con openpyxl para estilo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vista Operativa"

    # Estilos de encabezado y bordes
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Escribir Cabecera
    headers = list(df_operativo.columns)
    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Escribir Datos
    for row in df_operativo.itertuples(index=False):
        ws.append(list(row))

    # Formatear filas y celdas
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border_thin
            col_name = headers[col_idx - 1]

            # Formatos numéricos y alineación
            if col_name in ["Pedido", "PosPed", "Material", "N° Componentes"]:
                cell.alignment = Alignment(horizontal="center")
            elif col_name in ["Ctd.ped.", "Cantidad", "Cantidad_Total_Requerida"]:
                cell.alignment = Alignment(horizontal="right")
            elif col_name == "ValorNeto":
                cell.alignment = Alignment(horizontal="right")
            elif col_name == "Estado":
                cell.alignment = Alignment(horizontal="center")
                if cell.value == "REVISIÓN MATERIAL 37":
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                elif cell.value == "PRODUCTIVO":
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # Activar Filtro Automático en Excel
    ws.auto_filter.ref = ws.dimensions

    # Autoajustar ancho de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(salida_excel_operativo)
    print(f"¡Éxito! Vista Operativa de Excel generada en:\n{salida_excel_operativo}")


# --- PUNTO DE ENTRADA ---
if __name__ == "__main__":
    try:
        df_v, df_r = guardar_en_mysql()
        generar_excel_operativo(df_v, df_r)
    except Exception as e:
        print(f"\n[ERROR]: {e}")